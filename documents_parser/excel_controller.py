import re
from datetime import datetime
import xlrd
import openpyxl

def extract_text_from_excel(excel_path,tag,file_name ,user="Test_User"):

    extracted_data_exact = []
    extracted_data_partial = []

    def get_column_letter(column_index):
        return openpyxl.utils.get_column_letter(column_index + 1)

    if excel_path.endswith('.xls'):
        workbook = xlrd.open_workbook(excel_path)
        for sheet in workbook.sheets():
            sheet_name = sheet.name
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                for col_idx, cell in enumerate(row):
                    
                    cell_text = str(cell.value)
                    
                    exact_matches = list(re.finditer(rf"\b{re.escape(tag)}\b", cell_text, re.IGNORECASE))
                    for match in exact_matches:

                        cell_letter = get_column_letter(col_idx)
                        extracted_data_exact.append({
                            "Source File Name": f"{file_name}",
                            "File Type": 'Excel',
                            "Tag Searched": tag,
                            "Block/Record": cell_text,
                            "Location of the Tag": f"Sheet:{sheet_name}, Row: {row_idx + 1}, Cell: {cell_letter}",
                            "Date of Search": datetime.now().strftime("%B %d, %Y %H:%M:%S"),
                            "Search Author": user,
                            "Other": ""
                        })

                    # Find partial matches in the cell text
                    partial_matches = list(re.finditer(rf"\w*{re.escape(tag)}\w*", cell_text, re.IGNORECASE))
                    for match in partial_matches:
                        matched_text = match.group()
                        if matched_text.lower() != tag.lower():  # Avoid duplicates with exact matches

                            cell_letter = get_column_letter(col_idx)
                            extracted_data_partial.append({
                                "Source File Name": f"{file_name}",
                                "File Type": 'Excel',
                                "Tag Searched": tag,
                                "Block/Record": cell_text,
                                "Location of the Tag": f"Sheet:{sheet_name}, Row: {row_idx + 1}, Cell: {cell_letter}",
                                "Date of Search": datetime.now().strftime("%B %d, %Y %H:%M:%S"),
                                "Search Author": user,
                                "Other": "Partial match"
                            })

    elif excel_path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(excel_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for col_idx, cell in enumerate(row):
                    if cell is not None:
                        cell_text = str(cell)
                        
                        # Find exact matches in the cell text
                        exact_matches = list(re.finditer(rf"\b{re.escape(tag)}\b", cell_text, re.IGNORECASE))
                        for match in exact_matches:

                            cell_letter = get_column_letter(col_idx)
                            extracted_data_exact.append({
                                "Source File Name": f"{excel_path}",
                                "File Type": 'Excel',
                                "Tag Searched": tag,
                                "Block/Record": cell_text,
                                "Location of the Tag": f"Sheet:{sheet_name} Row: {row_idx}, Cell: {cell_letter}",
                                "Date of Search": datetime.now().strftime("%B %d, %Y %H:%M:%S"),
                                "Search Author": user,
                                "Other": ""
                            })

                        # Find partial matches in the cell text
                        partial_matches = list(re.finditer(rf"\w*{re.escape(tag)}\w*", cell_text, re.IGNORECASE))
                        for match in partial_matches:
                            matched_text = match.group()
                            if matched_text.lower() != tag.lower():  # Avoid duplicates with exact matches

                                cell_letter = get_column_letter(col_idx)
                                extracted_data_partial.append({
                                    "Source File Name": f"{excel_path}",
                                    "File Type": 'Excel',
                                    "Tag Searched": tag,
                                    "Block/Record": cell_text,
                                    "Location of the Tag": f" Sheet:{sheet_name} Row: {row_idx}, Cell: {cell_letter}",
                                    "Date of Search": datetime.now().strftime("%B %d, %Y %H:%M:%S"),
                                    "Search Author": user,
                                    "Other": "Partial match"
                                })
    return extracted_data_exact , extracted_data_partial





